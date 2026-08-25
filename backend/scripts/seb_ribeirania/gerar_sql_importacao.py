#!/usr/bin/env python3
"""Gera SQL idempotente de importação SEB Ribeirânia a partir das planilhas."""

from __future__ import annotations

import re
import unicodedata
from datetime import date, datetime, time
from pathlib import Path

from openpyxl import load_workbook

PASTA_XLSX = next(Path("/Users/lucasmoraes/Downloads").glob("DadosRibeir*2"))
ROOT = Path(__file__).resolve().parents[2]
SAIDA = ROOT / "database" / "migrations" / "2026_08_25_importar_dados_seb_ribeirania.sql"
SAIDA_ROLLBACK = ROOT / "database" / "migrations" / "2026_08_25_importar_dados_seb_ribeirania_rollback.sql"
SAIDA_COPIA = Path(__file__).resolve().parent / "importar_seb_ribeirania.sql"

SENHA_HASH = "$2y$12$cn8j9xCrBWl1F0fjzN0GM.F4htG2eYICiXj9SN76UvZCG3sZ3eQYG"
BANCO = "educatudo_seb_ribeirania"

CURSO_CANONICO = {
    "fundamental ii regular": "Fundamental II Regular",
    "ensino fundamental ii regular": "Fundamental II Regular",
    "ensino médio regular": "Ensino Médio Regular",
    "ensino medio regular": "Ensino Médio Regular",
    "fundamental ii bilingue": "Fundamental II Bilingue",
    "fundamental ii biblingue": "Fundamental II Bilingue",
    "fundamental ii bilingüe": "Fundamental II Bilingue",
}

COMPONENTE_CANONICO = {
    "lingua inglesa": "Língua Inglesa",
    "língua inglesa": "Língua Inglesa",
    "science": "Science/Steam",
    "science/steam": "Science/Steam",
    "biologia": "Biologia",
    "quimica": "Química",
    "química": "Química",
    "sociologia": "Sociologia",
    "lingua espanhola": "Língua Espanhola",
    "língua espanhola": "Língua Espanhola",
}

AREA_MAP = {
    "linguagens": "linguagens",
    "matematica": "matematica",
    "matemática": "matematica",
    "ciencias da natureza": "ciencias_natureza",
    "ciências da natureza": "ciencias_natureza",
    "ciencias humanas": "ciencias_humanas",
    "ciências humanas": "ciencias_humanas",
    "itinerario formativo": "outra",
    "itinerário formativo": "outra",
}

COR_AREA = {
    "linguagens": "#3B82F6",
    "matematica": "#F59E0B",
    "ciencias_natureza": "#10B981",
    "ciencias_humanas": "#EF4444",
    "outra": "#8B5CF6",
}

DIA_SEMANA = {
    "segunda": 1,
    "terca": 2,
    "terça": 2,
    "quarta": 3,
    "quinta": 4,
    "sexta": 5,
    "sabado": 6,
    "sábado": 6,
    "domingo": 7,
}

SALA_TURMA = {
    "5A": "Sala 5 A",
    "2C": "Sala 2 C",
}


def norm(s) -> str:
    if s is None:
        return ""
    t = unicodedata.normalize("NFC", str(s)).strip()
    t = re.sub(r"\s+", " ", t)
    return t


def sem_acento(s: str) -> str:
    nfkd = unicodedata.normalize("NFKD", s)
    return "".join(c for c in nfkd if not unicodedata.combining(c)).lower()


def chave(s) -> str:
    return sem_acento(norm(s))


def sql_str(v) -> str:
    if v is None:
        return "NULL"
    s = unicodedata.normalize("NFC", str(v))
    return "'" + s.replace("\\", "\\\\").replace("'", "''") + "'"


def sql_int(v, default="NULL") -> str:
    if v is None or v == "":
        return default
    return str(int(v))


def sql_date(v) -> str:
    if v is None:
        return "NULL"
    if isinstance(v, datetime):
        return sql_str(v.date().isoformat())
    if isinstance(v, date):
        return sql_str(v.isoformat())
    return sql_str(str(v)[:10])


def sql_time(v) -> str:
    if v is None:
        return "NULL"
    if isinstance(v, datetime):
        v = v.time()
    if isinstance(v, time):
        return sql_str(v.strftime("%H:%M:%S"))
    return sql_str(str(v))


def sim(v) -> int:
    return 0 if chave(v) in ("nao", "não", "n", "0", "false") else 1


def nonempty(row) -> bool:
    return any(c is not None and str(c).strip() != "" for c in row)


def sheet_rows(path: Path, sheet: str) -> list[dict]:
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb[sheet]
    rows = [r for r in ws.iter_rows(values_only=True) if nonempty(r)]
    wb.close()
    if not rows:
        return []
    headers = [norm(h) for h in rows[0]]
    out = []
    for row in rows[1:]:
        item = {}
        for i, h in enumerate(headers):
            item[h] = row[i] if i < len(row) else None
        out.append(item)
    return out


def xlsx(nome: str) -> Path:
    return PASTA_XLSX / nome


def curso_nome(v) -> str:
    k = chave(v)
    if k in CURSO_CANONICO:
        return CURSO_CANONICO[k]
    return norm(v)


def componente_nome(v) -> str:
    k = chave(v)
    if k in COMPONENTE_CANONICO:
        return COMPONENTE_CANONICO[k]
    return norm(v)


def area_enum(v) -> str:
    k = chave(v)
    return AREA_MAP.get(k, "outra")


def tipo_enum(nome: str, area_orig: str) -> str:
    k = chave(area_orig)
    n = chave(nome)
    if "itinerario" in k:
        return "itinerario_formativo"
    if n in ("lingua inglesa", "língua inglesa", "language", "lingua espanhola", "língua espanhola"):
        return "lingua_adicional"
    if n in ("coding",):
        return "complementar"
    return "formacao_geral"


def tipo_sala(nome: str) -> str:
    k = chave(nome)
    if any(x in k for x in ("lab", "explorat", "quimica", "fisica", "biologia", "tecnolog", "3d", "rv")):
        return "laboratorio"
    if "anf" in k:
        return "outro"
    if "sala de estudo" in k or "estudo ingles" in k:
        return "outro"
    return "sala"


def slug_codigo(*parts, maxlen=30) -> str:
    bits = []
    for p in parts:
        s = chave(p)
        s = re.sub(r"[^a-z0-9]+", "-", s).strip("-")
        if s:
            bits.append(s)
    code = "-".join(bits).upper()[:maxlen]
    return code or "SALA"


def main() -> None:
    escola = sheet_rows(xlsx("01_Dados_da_Escola.xlsx"), "Dados da Escola")[0]
    anos = sheet_rows(xlsx("02_Estrutura_Academica.xlsx"), "Ano Letivo")
    cursos = sheet_rows(xlsx("02_Estrutura_Academica.xlsx"), "Cursos")
    series = sheet_rows(xlsx("02_Estrutura_Academica.xlsx"), "Séries")
    componentes = sheet_rows(xlsx("03_Componentes_Curriculares.xlsx"), "Componentes Curriculares")
    matrizes = sheet_rows(xlsx("04_Matrizes_Curriculares.xlsx"), "Matrizes")
    matriz_comp = sheet_rows(xlsx("04_Matrizes_Curriculares.xlsx"), "Componentes da Matriz")
    salas = sheet_rows(xlsx("05_Salas_Ambientes.xlsx"), "Salas e Ambientes")
    turmas = sheet_rows(xlsx("06_Turmas.xlsx"), "Turmas")
    professores = sheet_rows(xlsx("07_Professores.xlsx"), "Professores")
    vinculos = sheet_rows(xlsx("08_Vinculos_Professores.xlsx"), "Vínculos Professores")
    grade = sheet_rows(xlsx("09_Grade_Horaria.xlsx"), "Grade Horária")
    alunos = sheet_rows(xlsx("10_Alunos.xlsx"), "Alunos")
    matriculas = sheet_rows(xlsx("11_Matriculas.xlsx"), "Matrículas")

    nomes_comp = {componente_nome(r["Nome"]) for r in componentes}
    for r in matriz_comp:
        n = componente_nome(r["Componente Curricular"])
        if n and n not in nomes_comp:
            componentes.append(
                {
                    "Nome": n,
                    "Código": None,
                    "Sigla": n[:10].upper(),
                    "Área do Conhecimento": "Itinerário Formativo",
                    "Descrição": None,
                    "Permite Avaliação": "Não",
                    "Controla Frequência": "Sim",
                    "Permite Diário": "Não",
                    "Ordem": None,
                    "Ativo": "Sim",
                }
            )
            nomes_comp.add(n)

    # Salas extras referenciadas nas turmas/grade
    nomes_sala = {(norm(r["Nome"]), norm(r["Bloco / Local"])) for r in salas}
    extras = [
        ("Sala 2 C", "C", 42, "2C"),
        ("Sala 2", "C", 50, "SALA-2"),
    ]
    for nome, bloco, cap, codigo in extras:
        if all(norm(s["Nome"]) != nome for s in salas):
            salas.append(
                {
                    "Nome": nome,
                    "Código": codigo,
                    "Tipo": "Sala",
                    "Capacidade": cap,
                    "Bloco / Local": bloco,
                    "Andar": None,
                    "Descrição": None,
                    "Ativa": "Sim",
                    "_codigo_fixo": codigo,
                }
            )

    linhas: list[str] = []
    a = linhas.append

    a("-- =============================================================================")
    a("-- Importação SEB Ribeirânia (tenant) — estrutura, turmas, alunos e grade")
    a("-- Arquivo UTF-8 (NFC). O PDO do Master já conecta com charset=utf8mb4.")
    a("-- Idempotente: INSERT ... WHERE NOT EXISTS.")
    a("-- Senha padrão alunos/professores: 123456 (primeiro acesso).")
    a("--")
    a("-- Master: /master/migrations → escola SEB → Escolher → marcar só este arquivo.")
    a("-- Não usar \"Executar todas\" (este arquivo é pulado no bootstrap de escola nova).")
    a("-- Rode antes as migrations de schema pendentes da escola.")
    a("-- =============================================================================")
    a("SET NAMES utf8mb4 COLLATE utf8mb4_unicode_ci;")
    a("SET character_set_client = utf8mb4;")
    a("SET character_set_connection = utf8mb4;")
    a("SET character_set_results = utf8mb4;")
    a("SET time_zone = '-03:00';")
    a("")
    a("-- Aborta se o banco atual não for da SEB (evita importar em outra escola).")
    a("SET @db_atual := DATABASE();")
    a("SET @eh_seb := (@db_atual LIKE '%seb%' OR @db_atual LIKE '%ribeirania%');")
    a("SET @stmt_guard := IF(@eh_seb, 'SELECT 1', CONCAT('SELECT * FROM `ERRO_importe_somente_na_SEB_banco_atual_e_', IFNULL(@db_atual, 'vazio'), '`'));")
    a("PREPARE stmt_guard FROM @stmt_guard;")
    a("EXECUTE stmt_guard;")
    a("DEALLOCATE PREPARE stmt_guard;")
    a("")

    # --- Unidade ---
    a("-- 1) Unidade (dados da escola)")
    a("INSERT INTO unidades (nome, tipo, razao_social, cnpj, endereco, numero, bairro, cidade, uf, cep, diretor_nome, secretario_nome, ativo)")
    a("SELECT {nome}, 'matriz', {razao}, {cnpj}, {end}, {num}, {bairro}, {cid}, {uf}, {cep}, {dir}, {sec}, 1 FROM DUAL".format(
        nome=sql_str(norm(escola["Nome Fantasia / Escola"])),
        razao=sql_str(norm(escola["Razão Social"]).replace("Brasilieiro", "Brasileiro")),
        cnpj=sql_str(norm(escola["CNPJ"])),
        end=sql_str(norm(escola["Logradouro"])),
        num=sql_str(norm(escola["Número"])),
        bairro=sql_str(norm(escola["Bairro"])),
        cid=sql_str(norm(escola["Cidade"])),
        uf=sql_str(norm(escola["UF"])),
        cep=sql_str(norm(escola["CEP"]).replace("-", "")),
        dir=sql_str(norm(escola["Diretor(a)"])),
        sec=sql_str(norm(escola["Secretário(a) Escolar"])),
    ))
    a("WHERE NOT EXISTS (SELECT 1 FROM unidades WHERE cnpj = {cnpj} OR nome = {nome});".format(
        cnpj=sql_str(norm(escola["CNPJ"])),
        nome=sql_str(norm(escola["Nome Fantasia / Escola"])),
    ))
    a("UPDATE unidades SET")
    a("  nome = {nome},".format(nome=sql_str(norm(escola["Nome Fantasia / Escola"]))))
    a("  razao_social = {razao},".format(razao=sql_str(norm(escola["Razão Social"]).replace("Brasilieiro", "Brasileiro"))))
    a("  cnpj = {cnpj},".format(cnpj=sql_str(norm(escola["CNPJ"]))))
    a("  endereco = {end},".format(end=sql_str(norm(escola["Logradouro"]))))
    a("  numero = {num},".format(num=sql_str(norm(escola["Número"]))))
    a("  bairro = {bairro},".format(bairro=sql_str(norm(escola["Bairro"]))))
    a("  cidade = {cid},".format(cid=sql_str(norm(escola["Cidade"]))))
    a("  uf = {uf},".format(uf=sql_str(norm(escola["UF"]))))
    a("  cep = {cep},".format(cep=sql_str(norm(escola["CEP"]))))
    a("  diretor_nome = {dir},".format(dir=sql_str(norm(escola["Diretor(a)"]))))
    a("  secretario_nome = {sec}".format(sec=sql_str(norm(escola["Secretário(a) Escolar"]))))
    a("WHERE id = (SELECT id FROM (SELECT MIN(id) AS id FROM unidades) x);")
    a("")

    # --- Ano letivo ---
    a("-- 2) Ano letivo")
    for r in anos:
        a("INSERT INTO ano_letivo (ano, data_inicio, data_fim, ativo)")
        a("SELECT {ano}, {ini}, {fim}, {ativo} FROM DUAL".format(
            ano=int(r["Ano"]),
            ini=sql_date(r["Data de Início"]),
            fim=sql_date(r["Data de Término"]),
            ativo=sim(r["Ativo"]),
        ))
        a("WHERE NOT EXISTS (SELECT 1 FROM ano_letivo WHERE ano = {ano});".format(ano=int(r["Ano"])))
    a("")

    # --- Cursos ---
    a("-- 3) Cursos (tabela curso da estrutura nova)")
    for i, r in enumerate(cursos, 1):
        nome = curso_nome(r["Nome"])
        a("INSERT INTO curso (nome, tipo, possui_serie, descricao, ativo, ordem)")
        a("SELECT {nome}, 'regular', 1, {desc}, {ativo}, {ordem} FROM DUAL".format(
            nome=sql_str(nome),
            desc=sql_str(norm(r.get("Descrição"))) if norm(r.get("Descrição")) else "NULL",
            ativo=sim(r.get("Ativo")),
            ordem=int(r["Ordem"]) if r.get("Ordem") not in (None, "") else i * 10,
        ))
        a("WHERE NOT EXISTS (SELECT 1 FROM curso WHERE LOWER(TRIM(nome)) = LOWER({nome}));".format(nome=sql_str(nome)))
    a("")

    # --- Séries ---
    a("-- 4) Séries")
    for i, r in enumerate(series, 1):
        curso = curso_nome(r["Curso"])
        nome = norm(r["Nome da Série/Ano"])
        a("INSERT INTO serie (curso_id, nome, ordem, ativo)")
        a("SELECT c.id, {nome}, {ordem}, {ativo} FROM curso c".format(
            nome=sql_str(nome),
            ordem=int(r["Ordem"]) if r.get("Ordem") not in (None, "") else i * 10,
            ativo=sim(r.get("Ativa")),
        ))
        a("WHERE c.nome = {curso}".format(curso=sql_str(curso)))
        a("  AND NOT EXISTS (SELECT 1 FROM serie s WHERE s.curso_id = c.id AND s.nome = {nome});".format(nome=sql_str(nome)))
    a("")

    # --- Componentes ---
    a("-- 5) Componentes curriculares (tabela materias)")
    a("--    Atualiza código/sigla se o catálogo padrão já tiver o mesmo nome.")
    for i, r in enumerate(componentes, 1):
        nome = componente_nome(r["Nome"])
        codigo = norm(r.get("Código")) or nome[:20].upper().replace(" ", "")
        sigla = (norm(r.get("Sigla")) or nome[:10].upper())[:10]
        area = area_enum(r.get("Área do Conhecimento"))
        tipo = tipo_enum(nome, r.get("Área do Conhecimento") or "")
        cor = COR_AREA.get(area, "#3B82F6")
        a("UPDATE materias SET codigo = {cod}, sigla = {sig}, area_conhecimento = {area}, tipo = {tipo},".format(
            cod=sql_str(str(codigo)[:20]),
            sig=sql_str(sigla),
            area=sql_str(area),
            tipo=sql_str(tipo),
        ))
        a("  permite_avaliacao = {av}, permite_frequencia = {fr}, permite_diario = {di}, ativo = {at}, ordem = {ord}".format(
            av=sim(r.get("Permite Avaliação")),
            fr=sim(r.get("Controla Frequência")),
            di=sim(r.get("Permite Diário")),
            at=sim(r.get("Ativo")),
            ord=int(r["Ordem"]) if r.get("Ordem") not in (None, "") else i,
        ))
        a("WHERE LOWER(TRIM(nome)) = LOWER({nome});".format(nome=sql_str(nome)))
        a("INSERT INTO materias (nome, codigo, sigla, area_conhecimento, tipo, descricao, cor, ordem,")
        a("  permite_avaliacao, permite_frequencia, permite_plano_aula, permite_diario, ativo,")
        a("  etapa_infantil, etapa_fund_i, etapa_fund_ii, etapa_medio)")
        a("SELECT {nome}, {cod}, {sig}, {area}, {tipo}, NULL, {cor}, {ord}, {av}, {fr}, 1, {di}, {at},".format(
            nome=sql_str(nome),
            cod=sql_str(str(codigo)[:20]),
            sig=sql_str(sigla),
            area=sql_str(area),
            tipo=sql_str(tipo),
            cor=sql_str(cor),
            ord=int(r["Ordem"]) if r.get("Ordem") not in (None, "") else i,
            av=sim(r.get("Permite Avaliação")),
            fr=sim(r.get("Controla Frequência")),
            di=sim(r.get("Permite Diário")),
            at=sim(r.get("Ativo")),
        ))
        a("  'nao_aplica', 'nao_aplica', 'obrigatoria', 'oferta' FROM DUAL")
        a("WHERE NOT EXISTS (SELECT 1 FROM materias WHERE LOWER(TRIM(nome)) = LOWER({nome}));".format(nome=sql_str(nome)))
        a("")

    # --- Matrizes ---
    a("-- 6) Matrizes curriculares")
    matriz_codigo = {
        "Ribeirânia 2026 - 8º Ano": "RIB-2026-8ANO",
        "Ribeirânia 2026 - 9º Ano": "RIB-2026-9ANO",
        "Ribeirânia 2026 - 8th": "RIB-2026-8TH",
        "Ribeirânia 2026 - 1ª Série": "RIB-2026-1SER",
        "Ribeirânia 2026 - 2ª Série": "RIB-2026-2SER",
    }
    ch_por_matriz: dict[str, float] = {}
    for r in matriz_comp:
        m = norm(r["Matriz"])
        try:
            ch_por_matriz[m] = ch_por_matriz.get(m, 0) + float(r.get("Carga Horária Anual (h)") or 0)
        except (TypeError, ValueError):
            pass

    for r in matrizes:
        nome = norm(r["Nome da Matriz"])
        codigo = matriz_codigo.get(nome, slug_codigo(nome))
        curso = curso_nome(r["Curso"])
        serie = norm(r["Série/Ano"])
        a("INSERT INTO matrizes_curriculares (nome, codigo, curso_id, serie_id, modalidade, turno,")
        a("  carga_horaria_anual_prevista, dias_letivos_previstos, duracao_padrao_aula_minutos, ativo)")
        a("SELECT {nome}, {cod}, c.id, s.id, NULL, {turno}, {ch}, {dias}, {dur}, {ativo}".format(
            nome=sql_str(nome),
            cod=sql_str(codigo),
            turno=sql_str(norm(r.get("Turno")) or "Matutino"),
            ch=ch_por_matriz.get(nome) or "NULL",
            dias=sql_int(r.get("Dias Letivos Previstos")),
            dur=sql_int(r.get("Duração Padrão da Aula (min)"), "45"),
            ativo=sim(r.get("Ativa")),
        ))
        a("FROM curso c")
        a("JOIN serie s ON s.curso_id = c.id AND s.nome = {serie}".format(serie=sql_str(serie)))
        a("WHERE c.nome = {curso}".format(curso=sql_str(curso)))
        a("  AND NOT EXISTS (SELECT 1 FROM matrizes_curriculares m WHERE m.codigo = {cod} OR m.nome = {nome});".format(
            cod=sql_str(codigo), nome=sql_str(nome)
        ))
        a("")

    a("-- 7) Componentes de cada matriz")
    for i, r in enumerate(matriz_comp, 1):
        matriz = norm(r["Matriz"])
        comp = componente_nome(r["Componente Curricular"])
        aulas = int(r.get("Aulas por Semana") or 1)
        obrig = sim(r.get("Obrigatório"))
        a("INSERT INTO matrizes_curriculares_componentes (matriz_id, materia_id, aulas_semana, obrigatorio, ordem_boletim, ordem_historico)")
        a("SELECT mx.id, mt.id, {aulas}, {ob}, {ord}, {ord}".format(aulas=aulas, ob=obrig, ord=i))
        a("FROM matrizes_curriculares mx")
        a("JOIN materias mt ON mt.id = (SELECT mt2.id FROM materias mt2 WHERE LOWER(TRIM(mt2.nome)) = LOWER({comp}) ORDER BY mt2.id LIMIT 1)".format(comp=sql_str(comp)))
        a("WHERE mx.nome = {matriz}".format(matriz=sql_str(matriz)))
        a("  AND NOT EXISTS (")
        a("    SELECT 1 FROM matrizes_curriculares_componentes x")
        a("    WHERE x.matriz_id = mx.id AND x.materia_id = mt.id")
        a("  );")
        a("")

    # --- Salas ---
    a("-- 8) Salas e ambientes")
    usados: set[str] = set()
    for i, r in enumerate(salas, 1):
        nome = norm(r["Nome"])
        bloco = norm(r.get("Bloco / Local"))
        codigo = norm(r.get("Código")) or r.get("_codigo_fixo") or slug_codigo(bloco, nome, str(i))
        if codigo in usados:
            codigo = f"{codigo}-{i}"[:30]
        usados.add(codigo)
        tipo = tipo_sala(nome)
        a("INSERT INTO school_locations (codigo, nome, tipo, capacidade, bloco, andar, ativo)")
        a("SELECT {cod}, {nome}, {tipo}, {cap}, {bloco}, {andar}, 1 FROM DUAL".format(
            cod=sql_str(codigo),
            nome=sql_str(nome),
            tipo=sql_str(tipo),
            cap=sql_int(r.get("Capacidade")),
            bloco=sql_str(bloco) if bloco else "NULL",
            andar=sql_str(norm(r.get("Andar"))) if norm(r.get("Andar")) else "NULL",
        ))
        a("WHERE NOT EXISTS (")
        a("  SELECT 1 FROM school_locations sl")
        a("  WHERE sl.codigo = {cod}".format(cod=sql_str(codigo)))
        a("     OR (sl.nome = {nome} AND IFNULL(sl.bloco,'') = IFNULL({bloco},''))".format(
            nome=sql_str(nome),
            bloco=sql_str(bloco) if bloco else "NULL",
        ))
        a(");")
    a("")

    # --- Turmas ---
    a("-- 9) Turmas")
    for r in turmas:
        nome = norm(r["Nome da Turma"])
        serie = norm(r["Série/Ano"])
        curso = curso_nome(r["Curso"])
        matriz = norm(r["Matriz Curricular"])
        sala_plan = norm(r.get("Sala Padrão"))
        sala_nome = SALA_TURMA.get(sala_plan, sala_plan)
        ano = int(r["Ano Letivo"])
        a("INSERT INTO turmas (nome, ano_letivo, ano_letivo_id, serie, serie_id, matriz_curricular_id,")
        a("  curso_novo_id, ativo, tipo_ensino, vagas, turno, sala_padrao_id, observacoes)")
        a("SELECT {nome}, {ano}, al.id, {serie}, s.id, mx.id, c.id, {ativo}, 'Ensino Fundamental II',".format(
            nome=sql_str(nome),
            ano=ano,
            serie=sql_str(serie),
            ativo=sim(r.get("Ativa")),
        ))
        a("  {vagas}, {turno}, sl.id, {obs}".format(
            vagas=sql_int(r.get("Capacidade")),
            turno=sql_str(norm(r.get("Turno")) or "Matutino"),
            obs=sql_str(norm(r.get("Observações"))) if norm(r.get("Observações")) else "NULL",
        ))
        a("FROM ano_letivo al")
        a("JOIN curso c ON c.nome = {curso}".format(curso=sql_str(curso)))
        a("JOIN serie s ON s.curso_id = c.id AND s.nome = {serie}".format(serie=sql_str(serie)))
        a("JOIN matrizes_curriculares mx ON mx.nome = {matriz}".format(matriz=sql_str(matriz)))
        a("LEFT JOIN school_locations sl ON sl.id = (")
        a("  SELECT sl2.id FROM school_locations sl2 WHERE sl2.nome = {sala} ORDER BY sl2.id LIMIT 1".format(sala=sql_str(sala_nome)))
        a(")")
        a("WHERE al.ano = {ano}".format(ano=ano))
        a("  AND NOT EXISTS (SELECT 1 FROM turmas t WHERE t.nome = {nome} AND t.ano_letivo = {ano});".format(
            nome=sql_str(nome), ano=ano
        ))
        a("")

    # --- Professores ---
    a("-- 10) Professores")
    a("INSERT INTO professores (nome, email, senha_hash, password, codigo_prof, materias, turmas, ativo, pagante)")
    a("SELECT 'A definir', NULL, {h}, '', 'ADEFINIR', JSON_ARRAY(), JSON_ARRAY(), 1, 0 FROM DUAL".format(h=sql_str(SENHA_HASH)))
    a("WHERE NOT EXISTS (SELECT 1 FROM professores WHERE codigo_prof = 'ADEFINIR');")
    a("")
    # professores table may not have primeiro_acesso - check educa_core... I didn't see primeiro_acesso on professores. Let me check.

    for r in professores:
        nome = norm(r["Nome Completo"])
        materias = sorted({componente_nome(v["Componente Curricular"]) for v in vinculos if norm(v["Professor"]) == nome})
        json_mat = "JSON_ARRAY(" + ", ".join(sql_str(m) for m in materias) + ")" if materias else "JSON_ARRAY()"
        a("INSERT INTO professores (nome, email, senha_hash, password, codigo_prof, materias, turmas, ativo, pagante)")
        a("SELECT {nome}, {email}, {h}, '', {cod}, {mat}, JSON_ARRAY(), 1, 1 FROM DUAL".format(
            nome=sql_str(nome),
            email=sql_str(norm(r.get("E-mail"))) if norm(r.get("E-mail")) else "NULL",
            h=sql_str(SENHA_HASH),
            cod=sql_str("PROF-" + re.sub(r"[^A-Z0-9]+", "", sem_acento(nome).upper())[:20]),
            mat=json_mat,
        ))
        a("WHERE NOT EXISTS (SELECT 1 FROM professores WHERE nome = {nome});".format(nome=sql_str(nome)))
        a("")

    a("-- Vínculo professor × turmas (JSON de IDs usado pelo painel do professor)")
    for r in professores:
        nome = norm(r["Nome Completo"])
        turmas_prof = sorted({norm(v["Turma"]) for v in vinculos if norm(v["Professor"]) == nome})
        if not turmas_prof:
            continue
        in_list = ", ".join(sql_str(t) for t in turmas_prof)
        a("UPDATE professores p SET p.turmas = (")
        a("  SELECT JSON_ARRAYAGG(t.id) FROM turmas t WHERE t.nome IN ({in_list}) AND t.ano_letivo = 2026".format(in_list=in_list))
        a(") WHERE p.nome = {nome};".format(nome=sql_str(nome)))
    a("")

    # --- Grade ---
    a("-- 11) Grade horária (pula horários sem componente; professor vazio → 'A definir')")
    vistos = set()
    for r in grade:
        turma = norm(r["Turma"])
        dia = DIA_SEMANA.get(chave(r["Dia da Semana"]))
        de = r["Horário Inicial"]
        ate = r["Horário Final"]
        comp = componente_nome(r.get("Componente Curricular"))
        if not turma or not dia or not de or not ate or not comp:
            continue
        key = (turma, dia, str(de), str(ate), comp)
        if key in vistos:
            continue
        vistos.add(key)
        prof = norm(r.get("Professor"))
        periodo = "manha" if chave(r.get("Turno")) in ("manha", "manhã", "matutino") else "tarde"
        a("INSERT INTO grade_horaria (dia_semana, horario_de, horario_ate, turma_id, professor_id, materia_id, periodo)")
        a("SELECT {dia}, {de}, {ate}, t.id, p.id, m.id, {per}".format(
            dia=dia,
            de=sql_time(de),
            ate=sql_time(ate),
            per=sql_str(periodo),
        ))
        a("FROM turmas t")
        a("JOIN materias m ON m.id = (SELECT m2.id FROM materias m2 WHERE LOWER(TRIM(m2.nome)) = LOWER({comp}) ORDER BY m2.id LIMIT 1)".format(comp=sql_str(comp)))
        if prof:
            a("JOIN professores p ON p.id = (SELECT p2.id FROM professores p2 WHERE p2.nome = {prof} ORDER BY p2.id LIMIT 1)".format(prof=sql_str(prof)))
        else:
            a("JOIN professores p ON p.codigo_prof = 'ADEFINIR'")
        a("WHERE t.id = (SELECT t2.id FROM turmas t2 WHERE t2.nome = {turma} AND t2.ano_letivo = 2026 ORDER BY t2.id LIMIT 1)".format(turma=sql_str(turma)))
        a("  AND NOT EXISTS (")
        a("    SELECT 1 FROM grade_horaria g")
        a("    WHERE g.turma_id = t.id AND g.dia_semana = {dia}".format(dia=dia))
        a("      AND g.horario_de = {de} AND g.horario_ate = {ate}".format(de=sql_time(de), ate=sql_time(ate)))
        a("      AND g.materia_id = m.id")
        a("  );")
        a("")

    # --- Alunos ---
    a("-- 12) Alunos (RA = número da matrícula da planilha 11)")
    a("SET @unidade_id := (SELECT MIN(id) FROM unidades);")
    mat_por_aluno = {norm(r["Aluno"]): r for r in matriculas}

    for r in alunos:
        nome = norm(r["Nome Completo"])
        m = mat_por_aluno.get(nome)
        if not m:
            raise SystemExit(f"Aluno sem matrícula: {nome}")
        ra = str(int(m["Número da Matrícula"])) if m.get("Número da Matrícula") not in (None, "") else None
        if not ra:
            raise SystemExit(f"Matrícula sem número: {nome}")
        serie = norm(m["Série/Ano"])
        turma = norm(m["Turma"])
        nick = "a" + ra
        sexo = norm(r.get("Sexo")).upper()[:1] or None
        if sexo not in ("M", "F", "N"):
            sexo = None
        a("INSERT INTO alunos (nome, nickname, email, senha_hash, password, ra, codigo_aluno, turma_id, unidade_id,")
        a("  serie, data_nasc, sexo, nacionalidade, ativo, status, pagante, primeiro_acesso)")
        a("SELECT {nome}, {nick}, NULL, {h}, '', {ra}, {ra}, t.id, @unidade_id,".format(
            nome=sql_str(nome),
            nick=sql_str(nick),
            h=sql_str(SENHA_HASH),
            ra=sql_str(ra),
        ))
        a("  {serie}, {nasc}, {sexo}, {nac}, 1, 'ACTIVE', 1, 1".format(
            serie=sql_str(serie),
            nasc=sql_date(r.get("Data de Nascimento")),
            sexo=sql_str(sexo) if sexo else "NULL",
            nac=sql_str(norm(r.get("Nacionalidade")) or "Brasileira"),
        ))
        a("FROM turmas t")
        a("WHERE t.id = (SELECT t2.id FROM turmas t2 WHERE t2.nome = {turma} AND t2.ano_letivo = 2026 ORDER BY t2.id LIMIT 1)".format(turma=sql_str(turma)))
        a("  AND NOT EXISTS (SELECT 1 FROM alunos a WHERE a.ra = {ra} OR a.nome = {nome});".format(
            ra=sql_str(ra), nome=sql_str(nome)
        ))
        a("")

    # --- Matrículas ---
    a("-- 13) Matrículas (aluno ↔ turma 2026)")
    for r in matriculas:
        nome = norm(r["Aluno"])
        ra = str(int(r["Número da Matrícula"]))
        turma = norm(r["Turma"])
        entrada = r.get("Data de Entrada") or r.get("Data da Matrícula")
        a("INSERT INTO matricula (aluno_id, turma_id, ano_letivo_id, data_entrada, status)")
        a("SELECT a.id, t.id, al.id, {entrada}, 'ativa'".format(entrada=sql_date(entrada)))
        a("FROM alunos a")
        a("JOIN turmas t ON t.id = (SELECT t2.id FROM turmas t2 WHERE t2.nome = {turma} AND t2.ano_letivo = 2026 ORDER BY t2.id LIMIT 1)".format(turma=sql_str(turma)))
        a("JOIN ano_letivo al ON al.ano = 2026")
        a("WHERE a.ra = {ra}".format(ra=sql_str(ra)))
        a("  AND NOT EXISTS (")
        a("    SELECT 1 FROM matricula m")
        a("    WHERE m.aluno_id = a.id AND m.turma_id = t.id AND m.ano_letivo_id = al.id")
        a("  );")
        a("")

    ras: list[str] = []
    for r in matriculas:
        ras.append(str(int(r["Número da Matrícula"])))
    ras_sql = ", ".join(sql_str(ra) for ra in ras)

    a("-- Conferência")
    a("SELECT 'ano_letivo' AS tabela, COUNT(*) AS qtd FROM ano_letivo")
    a("UNION ALL SELECT 'curso', COUNT(*) FROM curso")
    a("UNION ALL SELECT 'serie', COUNT(*) FROM serie")
    a("UNION ALL SELECT 'materias', COUNT(*) FROM materias")
    a("UNION ALL SELECT 'matrizes_curriculares', COUNT(*) FROM matrizes_curriculares")
    a("UNION ALL SELECT 'matrizes_curriculares_componentes', COUNT(*) FROM matrizes_curriculares_componentes")
    a("UNION ALL SELECT 'school_locations', COUNT(*) FROM school_locations")
    a("UNION ALL SELECT 'turmas', COUNT(*) FROM turmas")
    a("UNION ALL SELECT 'professores', COUNT(*) FROM professores")
    a("UNION ALL SELECT 'grade_horaria', COUNT(*) FROM grade_horaria")
    a("UNION ALL SELECT 'alunos', COUNT(*) FROM alunos")
    a("UNION ALL SELECT 'matricula', COUNT(*) FROM matricula;")
    a("")

    conteudo = "\n".join(linhas) + "\n"
    SAIDA.write_text(conteudo, encoding="utf-8")
    SAIDA_COPIA.write_text(conteudo, encoding="utf-8")

    rb: list[str] = []
    b = rb.append
    b("-- Rollback da importação SEB Ribeirânia (somente dados desta carga).")
    b("-- Não remove cursos, séries, componentes ou salas (podem ser reutilizados).")
    b("SET NAMES utf8mb4 COLLATE utf8mb4_unicode_ci;")
    b("DELETE FROM matricula WHERE aluno_id IN (SELECT id FROM alunos WHERE ra IN (" + ras_sql + "));")
    b("DELETE FROM alunos WHERE ra IN (" + ras_sql + ");")
    b("DELETE gh FROM grade_horaria gh INNER JOIN turmas t ON t.id = gh.turma_id WHERE t.nome IN ('5AM','2CM') AND t.ano_letivo = 2026;")
    b("DELETE FROM turmas WHERE nome IN ('5AM','2CM') AND ano_letivo = 2026;")
    b("DELETE FROM professores WHERE codigo_prof IN ('ADEFINIR','PROF-IGORZAPATADASILVA');")
    b("DELETE mcc FROM matrizes_curriculares_componentes mcc INNER JOIN matrizes_curriculares mx ON mx.id = mcc.matriz_id WHERE mx.codigo LIKE 'RIB-2026-%';")
    b("DELETE FROM matrizes_curriculares WHERE codigo LIKE 'RIB-2026-%';")
    b("")
    SAIDA_ROLLBACK.write_text("\n".join(rb) + "\n", encoding="utf-8")

    print(f"SQL gerado: {SAIDA} ({SAIDA.stat().st_size} bytes, {len(linhas)} linhas)")
    print(f"Rollback: {SAIDA_ROLLBACK}")
    print(f"  cursos={len(cursos)} series={len(series)} componentes={len(componentes)}")
    print(f"  matrizes={len(matrizes)} matriz_comp={len(matriz_comp)} salas={len(salas)}")
    print(f"  turmas={len(turmas)} professores={len(professores)} grade={len(vistos)}")
    print(f"  alunos={len(alunos)} matriculas={len(matriculas)}")


if __name__ == "__main__":
    main()
